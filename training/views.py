
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from users.models import Profile, Role
from .forms import (
    TrainingEventUpdateForm,
    UploadFileForm,
    TrainingModuleUpdateForm,
    NewTrainingEvent
)
from users.forms import UserRegisterForm, ProfileUpdateForm, UserRegisterForm2
from users.forms import RoleForm, RoleUpdateForm
from .models import TrainingEvent, TrainingModule, ProfileTrainingEvents, RoleTrainingModules, KPIValue
from django.contrib.auth.models import User
import pandas as pd
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime
import datetime as dt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from django.shortcuts import get_object_or_404



def home(request):
    return redirect('http://10.1.1.18')
    profile_instance = Profile.objects.get(user=request.user)
    must_have = profile_instance.must_have_training_modules()
    # go over must have training modules and check if they have been completed and if it's not save as '-
    data = []
    
    for training_module in must_have:
        event = TrainingEvent.objects.filter(profile=profile_instance, training_module=training_module).first()
        print(training_module, event)
        row = {}
        row['training_module'] = training_module
        row['event'] = event
        data.append(row) 
    print(data)
    percentage = profile_instance.get_training_modules_percentage()
   
    
    sidepanel = {
        'title': 'Training',
        'text1': 'Completed all trainings',
        'text2': 'Almost there',
    }
    
    context = {
        'title': 'Home',
        'sidepanel': sidepanel,
        'data': data,
        'percentage': percentage
    }
    return render(request, 'training/home.html', context)

@login_required
def percentage(request):
    profiles = Profile.objects.all()
    data = []
    for profile in profiles:
        row = {
            'username': profile.user.username,
            'percentage': profile.get_training_modules_percentage()
        }
        data.append(row)
    # sort them by percentage
    data = sorted(data, key=lambda x: x['percentage'], reverse=True)

    return render(request, 'training/percentage.html', {'title':'Percentage','data': data})


def supervisors(request):
    supervisors = User.objects.filter(supervisor_profiles__isnull=False, profile__active=True).distinct()
    training_modules = TrainingModule.objects.filter(other=False).order_by('name')
    users = User.objects.filter(profile__active=True).order_by('username')
    data = []
    if request.method == 'POST':
        current_supervisor_id = request.POST.get('current_supervisor')
        new_supervisor_id = request.POST.get('new_supervisor')
        print('POST:', request.POST)

        if current_supervisor_id and new_supervisor_id:
            current_supervisor = get_object_or_404(User, pk=current_supervisor_id)
            new_supervisor = get_object_or_404(User, pk=new_supervisor_id)

            # Update the supervisor for all relevant profiles
            profiles = current_supervisor.supervisor_profiles.all()
            for profile in profiles:
                profile.supervisor = new_supervisor
                profile.save()
            messages.success(request, f'Supervisor has been updated for {len(profiles)} profiles!')
    
    for sup in supervisors:
        
        print('Get supervisor:', sup.username)
        row = {
            'username': sup.first_name + ' ' + sup.last_name,
            'percentage':'',
            'total_modules': '',
            'modules':''
        }
        modules = {
            'completed': [],
            'expired': [],
            'missing': []
        }
        
        profiles = sup.supervisor_profiles.all()
        # filter only active profiles
        profiles = [profile for profile in profiles if profile.active]
        print('Profiles:', profiles)
        profile_training_events = ProfileTrainingEvents.objects.filter(profile__in=profiles).values_list('row', flat=True)
        
        for i, training_module in enumerate(training_modules):
            print('Training module:', training_module)
            for profile_training_event in profile_training_events:
                print('Profile training event:', profile_training_event)
                print('iii', i)
                profile_training_event = profile_training_event.split(',')[i]
                
                if profile_training_event == '-':
                    continue
                elif profile_training_event[0] == 'T':
                    modules['missing'].append(training_module)
                else:
                    try:
                        parsed_date = dt.datetime.strptime(profile_training_event, '%m/%d/%y')  # Assuming the date format is MM/DD/YY
                        current_date = dt.datetime.now()
                        delta = current_date - parsed_date
                        months_difference = delta.days // 30  # Approximate calculation for months
                        
                        if training_module.retrain_months:
                            if months_difference > training_module.retrain_months:
                                modules['expired'].append(training_module)
                            else:
                                modules['completed'].append(training_module)
                    except ValueError:
                        continue
        
        modules['completed'] = list(set(modules['completed']) - set(modules['expired']) - set(modules['missing']))
        modules['expired'] = list(set(modules['expired']))
        modules['missing'] = list(set(modules['missing']))
        
        # percentage of completed modules over total
        row['total_modules'] = len(modules['completed']) + len(modules['expired']) + len(modules['missing'])
        if row['total_modules'] == 0:
            row['total_modules'] = 1
        row['percentage'] = round(len(modules['completed']) / row['total_modules'] * 100)
        
        row['modules'] = modules
        data.append(row)
    # sort them by percentage
    data = sorted(data, key=lambda x: x['percentage'], reverse=True)
    
    return render(request, 'training/supervisors.html', {'title':'Supervisors','data': data, 'users': users, 'supervisors': supervisors})

@login_required
# def staff_roles(request):
#     profiles = Profile.objects.all()
#     sidepanel = {
#         'title': 'Training',
#         'text1': 'Completed all trainings',
#         'text2': 'Almost there',
#     }
#     context = {
#         'title': 'Staff Roles',
#         'profiles': profiles,
#         'sidepanel': sidepanel
#     }
#     return render(request, 'training/staff_roles.html', context)

# @login_required
# def roles(request):
#     roles = Role.objects.all()
#     sidepanel = {
#         'title': 'Training',
#         'text1': 'Completed all trainings',
#         'text2': 'Almost there',
#     }
#     context = {
#         'title': 'Roles',
#         'roles': roles,
#         'sidepanel': sidepanel
#     }
#     return render(request, 'training/roles.html', context)

@login_required
def new_entry(request):
    users = sorted(User.objects.filter(profile__active=True), key=lambda x: x.username)
    training_modules = sorted(TrainingModule.objects.all(), key=lambda x: x.name)
    if request.method == 'POST':
        user_ids = request.POST.getlist('user1')
        module_ids = request.POST.getlist('training_module1')
        print('POST:', request.POST)
        for user in user_ids:
            for module in module_ids:
                # create a new training event
                try:
                    training_event = TrainingEvent.objects.create(
                        profile=Profile.objects.get(user=user),
                        training_module=TrainingModule.objects.get(pk=module),
                        completed_date=request.POST['completed_date']
                    )
                    training_event.save()
                    print('Training event created:', training_event)
                    messages.success(request, f'Training event {training_event} has been added!')
                    
                except:
                    print('Error creating training event')
                    messages.error(request, 'Error creating training event')
        return redirect('training-history')
        
    form = NewTrainingEvent()
    sidepanel = {
        'title': 'Training',
        'text1': 'Completed all trainings',
        'text2': 'Almost there',
    }

    context = {
        'title': 'New Entry',
        'users': users,
        'training_modules': training_modules,
        'form': form,
        'sidepanel': sidepanel,
        'upload_form': UploadFileForm()
    }

    return render(request, 'training/new_entry.html', context)

@login_required
def new_user(request):
    if request.method == 'POST':
        # the request will receive only first name, last name and email we need to create a username and password with it
        username = request.POST['first_name'] + '_' + request.POST['last_name']
        # replace all spaces with underscores and make it lowercase
        username = username.replace(' ', '_').lower()
        password = username
        print('username:', username, 'password:', password)
        form_r = UserRegisterForm(data={'username': username, 'password1': password, 'password2': password, 'first_name': request.POST['first_name'], 'last_name': request.POST['last_name'], 'email': request.POST['email']}) 
        form_p = ProfileUpdateForm(request.POST)
        if form_r.is_valid() and form_p.is_valid():
            form_r.save()
            # update user profile with form_p data
            user = User.objects.get(username=form_r.cleaned_data['username'])
            form_p = ProfileUpdateForm(request.POST, instance=user.profile)
            form_p.save()
            profile_training_events = ProfileTrainingEvents.objects.create(profile=user.profile)
            profile_training_events.update_row()
            messages.success(request, 'New user has been created!')
            return redirect('training-new_user')
        else:
            messages.error(request, 'Form is not valid. Please check the entered data.')
    
    sidepanel = {
        'title': 'Training',
        'text1': 'Completed all trainings',
        'text2': 'Almost there',
    }

    context = {
        'title': 'New User',
        'form_u': UserRegisterForm2(),
        'form_p': ProfileUpdateForm(),
        'sidepanel': sidepanel
    }

    return render(request, 'training/new_user.html', context)

@login_required
def new_module(request):
    if request.method == 'POST':
        form = TrainingModuleUpdateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'New module has been created!')
            return redirect('training-grid')
        else:
            messages.error(request, 'Form is not valid. Please check the entered data.')

    else:
        form = TrainingModuleUpdateForm()
        return render(request, 'training/new_module.html', {'title':'New Module','form': form})
    
    sidepanel = {
        'title': 'Training',
        'text1': 'Completed all trainings',
        'text2': 'Almost there',
    }

    context = {
        'title': 'New User',
        'form_u': UserRegisterForm2(),
        'form_p': ProfileUpdateForm(),
        'sidepanel': sidepanel
    }

    return render(request, 'training/new_user.html', context)

def dashboard(request):
    # get kpi values for fully trained
    fully_trained = KPIValue.objects.filter(kpi__name='Percentage Fully Trained').order_by('date')
    training_performed = KPIValue.objects.filter(kpi__name='Training Performed').order_by('date')
    training_not_performed = KPIValue.objects.filter(kpi__name='Training Not Performed').order_by('date')
    retraining_not_performed = KPIValue.objects.filter(kpi__name='Retraining Not Performed').order_by('date')
    retraining_overdue = KPIValue.objects.filter(kpi__name='Retraining Overdue').order_by('date')
    # separate the values for a line graph
    fully_trained_values = [value.value for value in fully_trained]
    training_performed_values = [value.value for value in training_performed]
    
    training_not_performed_values = [value.value for value in training_not_performed]
    retraining_not_performed_values = [value.value for value in retraining_not_performed]
    retraining_overdue_values = [value.value for value in retraining_overdue]
    fully_trained_dates = [value.date.strftime('%Y-%m-%d') for value in fully_trained] 
    training_not_performed_dates = [value.date.strftime('%Y-%m-%d') for value in training_not_performed]
    training_performed_dates = [value.date.strftime('%Y-%m-%d') for value in training_performed]
    
    retraining_not_performed_dates = [value.date.strftime('%Y-%m-%d') for value in retraining_not_performed]
    retraining_overdue_dates = [value.date.strftime('%Y-%m-%d') for value in retraining_overdue]
    # how many training events have been completed this year, last year, etc
    profiles = Profile.objects.all()
    active_profiles = profiles.filter(active=True)
    training_events = TrainingEvent.objects.all()
    # only not other trainings
    training_modules = TrainingModule.objects.filter(other=False).order_by('name')
    # training events with certificate that have a retrain_months value
    # profiles_training_events = ProfileTrainingEvents.objects.filter(profile__in=profiles).select_related('profile')
    # this data is for total training and total retraining
    training = {
        'performed' : 0,
        'not_performed' : 0,
        'total' : 0
    }
    retraining = {
        'performed' : 0,
        'overdue' : 0,
        'not_performed' : 0,
        'total' : 0
    }
    profiles_fully_trained = 0
    training_performed_users = []
    training_not_performed_users=[]
    retraining_not_performed_users=[]
    retraining_overdue_users=[]
    for profile in active_profiles:
        try:
            profile_training_events = ProfileTrainingEvents.objects.filter(profile=profile).first()
            training_events_now = profile_training_events.row.split(',')
            fully_trained = True
        # combine trainin_modules and row.training_events in a for loop to check expired modules
            for i, training_module in enumerate(training_modules):
                    if training_events_now[i] == '-':
                        continue
                    elif training_events_now[i][0] == 'T':
                        fully_trained = False
                        if training_module.retrain_months:
                            retraining['total'] += 1
                            retraining['not_performed'] += 1
                            retraining_not_performed_users.append(profile.user)
                        else:
                            training['total'] += 1
                            training['not_performed'] += 1
                            training_not_performed_users.append(profile.user)
                        continue
        
                    # if it's a date compare it with the training_module retrain_months
                    try:
                        parsed_date = dt.datetime.strptime(training_events_now[i], '%m/%d/%y')  # Assuming the date format is MM/DD/YY
                        current_date = dt.datetime.now()
                        delta = current_date - parsed_date
                        months_difference = delta.days // 30  # Approximate calculation for months
                        
                        if training_module.retrain_months:
                            if months_difference > training_module.retrain_months:
                                retraining['overdue'] += 1
                                fully_trained = False
                                retraining_overdue_users.append(profile.user)
                            else:
                                retraining['performed'] += 1
                            retraining['total'] += 1

                        else:
                            training['performed'] += 1
                            training['total'] += 1
                            training_performed_users.append(profile.user)
                        
                            
                    except ValueError:
                        # do nothing
                        continue
        except:
            print('Error for:', profile.user.username)
            continue
        if fully_trained:
            profiles_fully_trained += 1

    perc_fully_trained = round(profiles_fully_trained / profiles.filter(active=True).count() * 100) if profiles.filter(active=True).count() else 0
    # remove duplicates from lists
    training_performed_users = list(set(training_performed_users))
    training_not_performed_users = list(set(training_not_performed_users))
    retraining_not_performed_users = list(set(retraining_not_performed_users))
    retraining_overdue_users = list(set(retraining_overdue_users))
    # round and divide by total to get the percentage of training and retraining
    training['performed'] = round(training['performed'] / training['total'] * 100) if training['total'] else 0
    retraining['performed'] = round(retraining['performed'] / retraining['total'] * 100) if retraining['total'] else 0
    retraining['overdue'] = round(retraining['overdue'] / retraining['total'] * 100) if retraining['total'] else 0
    training['not_performed'] = round(training['not_performed'] / training['total'] * 100) if training['total'] else 0
    retraining['not_performed'] = round(retraining['not_performed'] / retraining['total'] * 100) if retraining['total'] else 0
    
    # create data that counts how many profiles have been trained a year from now, between 1 and two and between 2 and five years)
    history = {
        '1year': 0,
        '2years': 0,
        '3years': 0,
        '5years': 0,
    }
    
    year3_users=[]
    year5_users=[]
    
    
    for profile in active_profiles:
        # year of the last training event
        event = TrainingEvent.objects.filter(profile=profile).order_by('-completed_date').last()
        # get the event that the completed date is the oldest

        if event:
            date = event.completed_date
            # convert date to timezone format
            date = datetime.combine(date, datetime.min.time())
            date = timezone.make_aware(date)
            if  timezone.now() - date < timezone.timedelta(days=365):
                history['1year'] += 1
            elif timezone.now() - date < timezone.timedelta(days=365*3):
                history['2years'] += 1
            elif timezone.now() - date < timezone.timedelta(days=365*5):
                history['3years'] += 1
                year3_users.append(profile.user)
            else:
                history['5years'] += 1
                year5_users.append(profile.user)
        else:
            print('No training event for:', profile.user.username)
    
    print('History:', history)
    history2 = {'x': ['1 year', '2 years', '3 years', '5 years'], 'y': [round(history['1year']/active_profiles.count()*100), round(history['2years']/active_profiles.count()*100), round(history['3years']/active_profiles.count()*100), round(history['5years']/active_profiles.count()*100)]}
    history1 = sorted(history.items(), key=lambda x: x[0])
    by_year = {}
    for event in training_events:
        year = event.completed_date.year
        if year in by_year:
            by_year[year] += 1
        else:
            by_year[year] = 1
    by_year1 = sorted(by_year.items(), key=lambda x: x[0])
    # reverse by_year1
    by_year1 = by_year1[::-1]
    # prepare this data for a chart
    print('By year1:', by_year1)

    by_year2 = {'x': [x[0] for x in by_year1], 'y': [x[1] for x in by_year1]}
    
    print('training_performed_users:', training_performed_values)
    print('training_performed_dates:', training_performed_dates)
    print('training_not_performed_values:', training_not_performed_values)
    print('training_not_performed_dates:', training_not_performed_dates)
    print('retraining_not_performed_values:', retraining_not_performed_values)
    print('retraining_not_performed_dates:', retraining_not_performed_dates)
    print('retraining_overdue_values:', retraining_overdue_values)
    print('retraining_overdue_dates:', retraining_overdue_dates)

    context = {
        'title': 'Dashboard',
        'profiles_count': profiles.filter(active=True).count(),
        'perc_fully_trained': perc_fully_trained,
        'by_year1':by_year1,
        'by_year2':by_year2,
        'history1': history1,
        'history2': history2,
        'training': training,
        'retraining': retraining,
        'training_not_performed_users': training_not_performed_users,
        'retraining_not_performed_users': retraining_not_performed_users,
        'retraining_overdue_users': retraining_overdue_users,
        'year3_users': year3_users,
        'year5_users': year5_users,
        'fully_trained_values': fully_trained_values,
        'fully_trained_dates': fully_trained_dates,
        'training_performed_values': training_performed_values,
        'training_performed_dates': training_performed_dates,
        'training_not_performed_values': training_not_performed_values,
        'training_not_performed_dates': training_not_performed_dates,
        'retraining_not_performed_values': retraining_not_performed_values,
        'retraining_not_performed_dates': retraining_not_performed_dates,
        'retraining_overdue_values': retraining_overdue_values,
        'retraining_overdue_dates': retraining_overdue_dates,
    }

    return render(request, 'training/dashboard.html', context)

def inactive(request):
    # get profiles with active = False
    inactive_users = Profile.objects.filter(active=False)
    sidepanel = {
        'title': 'Training',
        'text1': 'Completed all trainings',
        'text2': 'Almost there',
    }
    context = {
        'title': 'Inactive',
        'inactive_users': inactive_users,
        'sidepanel': sidepanel
    }
    return render(request, 'training/inactive.html', context)

@login_required
def history(request):

    selected_user = request.GET.get('user', '')
    users = User.objects.all().order_by('username')

    if selected_user:
        filter_user = User.objects.get(id=selected_user)
        training_events = TrainingEvent.objects.filter(profile=filter_user.profile).order_by('-created_date')
        
    else:
        training_events = TrainingEvent.objects.all().order_by('-created_date')
    paginator = Paginator(training_events, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    sidepanel = {
        'title': 'Training',
        'text1': 'Completed all trainings',
        'text2': 'Almost there',
    }
    context = {
        'title': 'History',
        'training_events': page_obj,
        'sidepanel': sidepanel,
        'users': users,
        'selected_user': int(selected_user) if selected_user else selected_user,
    }
    return render(request, 'training/history.html', context)

@login_required
def training_profile(request, profile_id):
    profile = Profile.objects.get(pk=profile_id)
    training_events = TrainingEvent.objects.filter(profile=profile).order_by('-completed_date')
    training_modules = profile.must_have_training_modules()
    print('Training modules:', training_modules)
    # zip training events with training modules
    data = []
    for training_module in training_modules:
        events = TrainingEvent.objects.filter(profile=profile, training_module=training_module).order_by('-completed_date')
        row = {}
        row['training_module'] = training_module
        row['events'] = events if events.exists() else 'missing'
        data.append(row)

    if request.method == 'POST':
        p_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile)
        if p_form.is_valid():
            p_form.save()
            messages.success(request,f'Account has been updated!')
            profile_training_events = ProfileTrainingEvents.objects.filter(profile=profile).first()
            profile_training_events.update_row()
            return redirect('training-profile', profile_id=profile_id)
    else:
        p_form = ProfileUpdateForm( instance= profile)

    sidepanel = {
        'title': 'Required Modules',
        'text1': 'Completed all trainings',
        'text2': '',
    }
    context = {
        'title': 'Profile',
        'profile': profile,
        'training_modules': training_modules,
        'training_events': training_events,
        'sidepanel': sidepanel,
        'p_form':p_form,
        'data': data
    }
    return render(request, 'users/profile.html', context)

def grid(request):
    # get all the supervisors
    supervisors = User.objects.filter(supervisor_profiles__isnull=False, profile__active=True).distinct()
    
    # if the request has a supervisor parameter, filter the profiles by the supervisor
    selected_supervisor = request.GET.get('supervisor', '')
    other = request.GET.get('other', '')
    print('request:', request.GET)

    if selected_supervisor:
        #  filter profile objects where supervisor = supervisor
        profiles = Profile.objects.filter(supervisor=selected_supervisor, active=True)
        print('Profiles:', profiles)
    else:
        profiles = Profile.objects.filter(active=True)

    # if other is false filter trainingmodules that have other false
    if other:
        training_modules = TrainingModule.objects.all().order_by('name')
    else:    
        training_modules = TrainingModule.objects.all().order_by('name').filter(other=False)
        out_of_grid = TrainingModule.objects.all().count() - TrainingModule.objects.all().filter(other=False).count()
        print('Out of grid:', out_of_grid)

    # Prepare data to be sent to the template
    data = []
    profile_training_events = ProfileTrainingEvents.objects.filter(profile__in=profiles).select_related('profile')
    for profile_training_event in profile_training_events:
        profile = profile_training_event.profile
        row = {
            'profile': profile,
            'roles': profile.roles.all(),
            'training_events': profile_training_event.row.split(',') if profile_training_event.row else [],
        }
        if other == '':
            # remove the items from training_events list, the last ones and the quantity is given by out_of_grid
            row['training_events'] = row['training_events'][:-(out_of_grid)]
        # combine trainin_modules and row.training_events in a for loop to check expired modules
        for i, training_module in enumerate(training_modules):
            # if list out of range add an extra element (This is for when adding a new module)
            if len(row['training_events']) < len(training_modules):
                row['training_events'].append('')
            if row['training_events'][i] == '-' or row['training_events'][i] == '' or row['training_events'][i] == '+':
                continue
            if row['training_events'][i][0] == 'T':
                continue
            # if it's a date compare it with the training_module retrain_months
            try:
                parsed_date = dt.datetime.strptime(row['training_events'][i], '%m/%d/%y')  # Assuming the date format is MM/DD/YY
                current_date = dt.datetime.now()
                delta = current_date - parsed_date
                months_difference = delta.days // 30  # Approximate calculation for months
                
                if training_module.retrain_months:
                    if months_difference > training_module.retrain_months:
                        row['training_events'][i] = 'Expired'
                        
                    elif months_difference > training_module.retrain_months - 4 and training_module.retrain_months > months_difference:
                        row['training_events'][i] = 'To Expire'
                    
                   
            except ValueError:
                print('this is the value that couldnt convert to date',row['training_events'][i])
                

        data.append(row)
    # order by the first name of the profile
    data = sorted(data, key=lambda x: x['profile'].user.first_name)

    
        
    # create a function that returns all the roles in a form format to display it 
    data2 = [
        {
            'role': obj.role,
            'training_modules': obj.row.split(',')
        }
        for obj in RoleTrainingModules.objects.all().order_by('role')
    ]
    # if other is false remove th last modules from the list
    if other == '':
        for obj in data2:
            obj['training_modules'] = obj['training_modules'][:-out_of_grid]


    # Check if download parameter is true
    download = request.GET.get('download', 'false').lower() == 'true'
    if download:
        # Create an Excel workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Training Data"

        # Add headers
        headers = ["First name", "Last name"] + [tm.name for tm in training_modules]
        ws.append(headers)

        print('Headers:', headers)

        # Apply styles to the header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Add data rows
        for item in data:
            profile = item['profile']
            first_name = profile.user.first_name
            last_name = profile.user.last_name
            training_events = item['training_events'] 

            row = [first_name, last_name] + training_events
            ws.append(row)

        # Apply styles to the data rows
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=len(headers), max_row=len(data) + 1):
            for cell in row:
                # if the first letter is "TM" then change to Missing
                if cell.value[0]== 'T' and cell.value[1] == 'M':
                    cell.value = 'Missing'
                if cell.value == 'Missing':
                    cell.font = Font(color="A9A9A9")  # Gray text
                elif cell.value == 'Expired':
                    cell.font = Font(color="FF0000")  # Red text
                elif cell.value == 'To Expire':
                    cell.font = Font(color="FFD700")  # Yellow text
                elif cell.value == '-':
                
                    cell.font = Font(color="000000")
                # else green text
                else:
                    cell.font = Font(color="008000")  

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        # Create an HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=training_data.xlsx'
        wb.save(response)
        return response

    context = {
        'title': 'Grid',
        'profiles': profiles,
        'training_modules': training_modules,
        'supervisors': supervisors,
        'selected_supervisor': int(selected_supervisor) if selected_supervisor else selected_supervisor,
        # if it was selected return true if it was not return false
        'selected_other': 'other' if other else '',
        'data': data,
        'data2': data2
    }

    return render(request, 'training/grid.html', context)


@login_required
def send_reminder_email(request, training_module_id):
    training_module = TrainingModule.objects.get(pk=training_module_id)
    certStatuses = training_module.get_incomplete_training_events()
    for certStatus in certStatuses:
        certStatus.send_schedule_notification()
        certStatus.save()
    messages.success(request, f'Reminder emails have been sent for {training_module.name} certificate!')



    return redirect('training-all_trainings')

@login_required
def training_event_detail(request, training_event_id):
    profiles = Profile.objects.all()
    training_event = TrainingEvent.objects.get(pk=training_event_id)
    form = TrainingEventUpdateForm(instance=training_event)
    if request.method == 'POST':
        form = TrainingEventUpdateForm(request.POST, instance=training_event)
        if form.is_valid():
            form.save()
            messages.success(request, f'{training_event.training_module} certificate has been updated!')
            return redirect('training-event-detail', training_event_id=training_event_id)
        else:
            messages.error(request, 'Form is not valid. Please check the entered data.')
    sidepanel = {
        'title': 'Training',
        'text1': 'Completed all trainings',
        'text2': 'Almost there',
    }

    context = {
        'title': training_event.training_module,
        'training_event': training_event,
        'profiles': profiles,
        'form': form,
        'sidepanel': sidepanel
    }
    return render(request, 'training/event_detail.html', context)

@login_required
def training_confirm_delete(request, training_event_id):
    training_event = TrainingEvent.objects.get(pk=training_event_id)
    context = {
        'title': 'Delete',
        'training_event': training_event,
        'event': training_event,
    }
    return render(request, 'training/confirm_delete.html', context)

@login_required
def training_delete(request, training_event_id):
    training_event = TrainingEvent.objects.get(pk=training_event_id)
    training_event.delete()
    messages.success(request, f'{training_event} certificate has been deleted!')
    return redirect('training-history')

@login_required
def training_event_delete(request, training_event_id):
    # are you sure you want to delete this training event?
    training_event = TrainingEvent.objects.get(pk=training_event_id)
    training_event.delete()
    messages.success(request, f'{training_event.training_module} certificate has been deleted!')
    return redirect('training-history')

@login_required
def training_role_detail(request, role_id):
    role = Role.objects.get(pk=role_id)
    role_modules = RoleTrainingModules.objects.filter(role=role).first()
    training_profiles = ProfileTrainingEvents.objects.all()
    form = RoleUpdateForm(instance=role)
    if request.method == 'POST':
        form = RoleUpdateForm(request.POST, instance=role)
        if form.is_valid():
            form.save()
            messages.success(request, f'{role.name} role has been updated!')
            role_modules.update_row()
            for profile in training_profiles:
                profile.update_row()
            return redirect('training-role-detail', role_id=role_id)
        else:
            messages.error(request, 'Form is not valid. Please check the entered data.')
    sidepanel = {
        'title': 'Training',
        'text1': 'Completed all trainings',
        'text2': 'Almost there',
    }

    context = {
        'title': role.name,
        'role': role,
        'form': form,
        'sidepanel': sidepanel
    }
    return render(request, 'training/role_detail.html', context)

@login_required
def training_module_detail(request, training_module_id):
    profiles = Profile.objects.all()
    training_module = TrainingModule.objects.get(pk=training_module_id)
    form = TrainingModuleUpdateForm(instance=training_module)
    if request.method == 'POST':
        form = TrainingModuleUpdateForm(request.POST, instance=training_module)
        if form.is_valid():
            form.save()
            messages.success(request, f'{training_module.name} certificate has been updated!')
            return redirect('training-module-detail', training_module_id=training_module_id)
        else:
            messages.error(request, 'Form is not valid. Please check the entered data.')
    sidepanel = {
        'title': 'Training',
        'text1': 'Completed all trainings',
        'text2': 'Almost there',
    }

    context = {
        'title': training_module.name,
        'training_module': training_module,
        'profiles': profiles,
        'form': form,
        'sidepanel': sidepanel
    }
    return render(request, 'training/module_detail.html', context)

@login_required
def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            
            # Rest of your code here
            excel_file = request.FILES['file']
            df = pd.read_excel(excel_file, sheet_name='FILL HERE_Performed Date')
            
            # save the values of the first five character of each header in a list
            training_modules = df.columns[1:]
            
            #  save the values of the header in a list
            
            for index, row in df.iterrows():
                # Rest of your code here
                employee_username = row['Employee'].replace(' ', '_').lower()  # Assuming 'Employee' is the column name in Excel
                
                print('Analyzing user:', employee_username)
                # Iterate through each certificate and its completion date
                for certificate in training_modules:
                    certificate_name = certificate[0:5]  # Get the first 5 characters of the certificate name
                    completion_date = row[certificate]
                    
                    if pd.notna(completion_date) and completion_date != 'R':
                        try:
                            # Retrieve the profile object using the username
                            user = User.objects.get(username=employee_username)
                            profile = Profile.objects.get(user=user)
                            training_module = TrainingModule.objects.get(name=certificate_name)

                            # Create TrainingEvent object
                            if not TrainingEvent.objects.filter(profile=profile, training_module=training_module, completed_date=completion_date).exists():
                                TrainingEvent.objects.create(
                                    profile=profile,
                                    training_module=training_module,
                                    completed_date=completion_date
                                )

                                print(f'TrainingEvent object created for {profile} and {training_module}')

                        except User.DoesNotExist:
                            print(f"User does not exist for {employee_username}")
                            continue  # Skip the iteration if the user doesn't exist
                        except (Profile.DoesNotExist, TrainingModule.DoesNotExist):
                            print(f"Profile or training_module does not exist for {employee_username} or {certificate_name}")
                            continue  # Skip the iteration if the profile or training_module doesn't exist
                    
                        
                # return render dashoboard.html with successful message
            messages.success(request, f'File has been uploaded successfully!')
            return redirect('training-history')
        else:
            messages.error(request, 'Form is not valid. Please check your input.')
    else:
        form = UploadFileForm()
    return render(request, 'upload_file.html', {'form': form})

# API routes

# class TrainingModules(APIView): 
#     permission_classes = [DjangoModelPermissionsOrAnonReadOnly]

#     def get_queryset(self):
#         return TrainingModule.objects.all().order_by('name')

#     def get(self, request):
#         training_modules = self.get_queryset()
#         serializer = TrainingModuleSerializer(training_modules, many=True)
#         return Response(serializer.data)

# class TrainingEvents(APIView):
#     permission_classes = [DjangoModelPermissionsOrAnonReadOnly]

#     def get_queryset(self):
#         return TrainingEvent.objects.all()

#     def get(self, request):
#         training_modules = self.get_queryset()
#         serializer = TrainingEventSerializer(training_modules, many=True)
#         return Response(serializer.data)

